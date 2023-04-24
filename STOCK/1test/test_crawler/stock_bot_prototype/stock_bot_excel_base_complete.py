from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

import pandas as pd
import requests
from bs4 import BeautifulSoup
import re




class saveNload:
    def __init__(self):
        pass

    def load(self):
        wb = load_workbook('result.xlsx')
        ws = wb['Sheet1']
        f = pd.DataFrame(ws.values)
        return f
    
    def save(self,df):
        wb = openpyxl.Workbook()
        # 시트 선택
        ws = wb.active
        # 시트에 데이터프레임 삽입
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        # 엑셀 파일에 저장
        wb.save("./result.xlsx")
        pass

class Main:
    def __init__(self):
        self.ls_tag=[]
        self.ls_price=[]
        self.url = "https://finance.naver.com/item/main.nhn?code="

        #태그 가져오기
        self.df_m = saveNload()
        self.df = self.df_m.load()
        self.ls_tag = self.df.iloc[:,0]
        print(self.ls_tag)
        del self.ls_tag[0]
                
    def get_price_one(self,tag):
        urltag = self.url + tag
        html = requests.get(urltag)
        soup = BeautifulSoup(html.content, 'html.parser')
        html1 = soup.find('div',class_='today')
        html2 = html1.find('p',class_='no_today')
        price = html2.find_all('em',class_='no_down')
        price = html2.find_all('span',class_='blind')
        result_price = re.sub(r'[^0-9]', '', str(price))
        print("price : ",result_price)
        return result_price

    def get_amount_one(self,tag):
        urltag = self.url + tag
        response = requests.get(urltag)
        html = response.text
        soup = BeautifulSoup(html,'html.parser')
        amount = soup.select_one('#tab_con1 > div.first > table > tr:nth-child(4) > td > em')
        rsult = amount.get_text().replace(',','')
        print("amount : ",rsult)
        return rsult

    def get_total_one(self,tag):
        urltag = self.url + tag
        response = requests.get(urltag)
        html = response.text
        soup = BeautifulSoup(html,'html.parser')
        amount = soup.select_one('#_market_sum')
        rsult = amount.get_text().replace(',','')
        return rsult

    def get_name(self,tag):
        urltag = self.url + tag
        response = requests.get(urltag)
        html = response.text
        soup = BeautifulSoup(html,'html.parser')
        amount = soup.select_one('#middle > div.h_company > div.wrap_company > h2 > a')
        rsult = amount.get_text()
        print("name : ",rsult)
        return rsult

    
    def update_price(self):
        count = 1
        for each in self.ls_tag:
            code = str(each)
            for i in range(6-len(code)):
                code = "0"+code
            price = self.get_price_one(code)
            self.df.loc[count,2] = int(price)

            amount = self.get_amount_one(code)
            self.df.loc[count,3] = int(amount)
            
            total = self.get_total_one(code)          
            self.df.loc[count,4] = total

            name = self.get_name(code)
            self.df.loc[count,1] = name
            count+=1

        print(self.df)

        
        self.df.to_excel("result.xlsx",header=None,index=False)
        


m = Main()
rsult = m.get_price_one('005930')
rsult = m.update_price()
