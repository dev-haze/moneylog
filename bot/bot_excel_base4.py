from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

import pandas as pd
import requests
from bs4 import BeautifulSoup
import re



wb = load_workbook('result.xlsx')
ws = wb['Sheet1']
 
f = pd.DataFrame(ws.values)

print(f)



class saveNload:
    def __init__(self):
        pass

    def load(self):
        wb = load_workbook('result.xlsx')
        ws = wb['Sheet1']
        f = pd.DataFrame(ws.values)
        return f
    
    def save(self,df):
        wb = openpyxl.Workbook()
        # 시트 선택
        ws = wb.active
        # 시트에 데이터프레임 삽입
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        # 엑셀 파일에 저장
        wb.save("./result.xlsx")
        pass

class Main:
    def __init__(self):
        self.ls_tag=[]
        self.ls_price=[]
        self.url = "https://finance.naver.com/item/main.nhn?code="

        #태그 가져오기
        self.df_m = saveNload()
        self.df = self.df_m.load()
        self.ls_tag = self.df[0]
        
    def get_price_one(self,tag):
        urltag = self.url + tag
        html = requests.get(urltag)
        soup = BeautifulSoup(html.content, 'html.parser')
        html1 = soup.find('div',class_='today')
        html2 = html1.find('p',class_='no_today')
        price = html2.find_all('em',class_='no_down')
        price = html2.find_all('span',class_='blind')
        result_price = re.sub(r'[^0-9]', '', str(price))
        return result_price

    def update_price(self):

        count = 1
        for each in self.ls_tag:
            try:
                price = self.get_price_one(each)
                print(price)
                self.df.loc[count,2] = price
            except:
                pass
            
        print(self.df)
        self.df.to_excel("result.xlsx",header=None,index=False)
        




m = Main()
rsult = m.get_price_one('005930')
rsult = m.update_price()
print(rsult)
